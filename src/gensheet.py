from datetime import datetime
import itertools
from openpyxl import Workbook
from openpyxl.styles import Font

# Below is excel sheet generation code
# it takes an argument that looks like [{'lb'}, {'top10', 'dh5a'}, {'repressilator'}, {'atc'}]
def excel_sheet(lists, replicates):
    product = list(itertools.product(*lists))
    wb = Workbook()
    sheet1 = wb.active
    sheet1.title = "Sample Design"
    titles = [
        "Sample Design ID",
        "Media ID",
        "Strain ID",
        "Vector ID",
        "Supplement ID",
    ]

    # Populate the first row with the titles
    for col_num, title in enumerate(titles, start=1):
        cell = sheet1.cell(row=1, column=col_num, value=title)
        cell.font = Font(bold=True)

    global sample_design_id
    sample_design_id = 1

    # Populating the excel sheet with generated data
    for row_num, item in enumerate(product, start=2):
        # Sample ID counter for first row
        sheet1.cell(row=row_num, column=1, value=f"SampleDesign{sample_design_id}")
        sample_design_id += 1

        # Continue to populate the rest of the row
        for col_num, value in enumerate(item, start=2):
            sheet1.cell(row=row_num, column=col_num, value=value)

    plate_96_wells_coordinates = [
        (row, col) for row in range(1, 9) for col in range(1, 13)
    ]

    def generate_sample_names(num_designs):
        global sample_names
        sample_names = []
        for i in range(0, sample_design_id - 1):
            sample_names.append(f"SampleDesign{i+1}")
        return sample_names

    num_designs = 4
    generate_sample_names(num_designs)

    # This creates/formats sheet 2 with the samples and their positions in a well plate
    sheet2 = wb.create_sheet(title="Sample")

    headers = ["Sample ID", "Row", "Column", "Well ID", "Sample Design ID"]

    for col_num, header in enumerate(headers, start=1):
        sheet2.cell(row=1, column=col_num, value=header)

    # Initializes the well count and sample ID count
    well_count = 0
    sample_id = 1

    for sd in sample_names:
        for r in range(replicates):
            row_data = [
                f"Sample{sample_id}",
                plate_96_wells_coordinates[well_count][0],
                plate_96_wells_coordinates[well_count][1],
                f"Assay{+1}",
                sd,
            ]
            sheet2.append(row_data)
            well_count += 1
            sample_id += 1

    # Auto fit column width code
    for col in sheet1.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        sheet1.column_dimensions[column].width = adjusted_width

    for col in sheet2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        sheet2.column_dimensions[column].width = adjusted_width

    timestamp = datetime.now().strftime("%Y-%m-%d_%H:%M:%S")
    filename = f"output/SampleDesign_{timestamp}.xlsx"

    wb.save(filename)
