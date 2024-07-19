from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from datetime import datetime

import itertools

from tkinter import *
from tkinter.ttk import *
from tkinter import Entry, Button, messagebox
import tkinter as tk
from tkinter import ttk

from PIL import Image, ImageTk

import ctypes

import os
import string
uppercase_alphabet = [chr(i) for i in range(ord('A'), ord('Z') + 1)]

#If the user wishes to add more metadata options, they can do so here
input_metadata_test = {
                'media':['lb', 'm9'],
                'strain':['top10', 'dh5a', 'ecoli'],
                'supplement':['atc', 'iptg'],
                'vector':['repressilator','toggleswitch'],
                }

#GUI format setup code below
root = tk.Tk()
root.title("LabGen Sample Design ID Generator")
root.geometry("600x550")


title_label = tk.Label(root, text = "SAMPLE ID SELECTION MENU", 
                       font = ("Arial", 12, "bold"), 
                       fg = "black",
                       )
title_label.grid(column =1, row =1, sticky=tk.N+tk.S+tk.W+tk.E)


hseperator1 = ttk.Separator(root, orient='horizontal')
hseperator1.grid(column=0, row=0, columnspan=7, sticky='ew', pady=10)

hseperator2 = ttk.Separator(root, orient='horizontal')
hseperator2.grid(column=0, row=5, columnspan=7, sticky='ew', pady=10)

hseperator3 = ttk.Separator(root, orient='horizontal')
hseperator3.grid(column=0, row=9, columnspan=7, sticky='ew', pady=10)

hseperator4 = ttk.Separator(root, orient='horizontal')
hseperator4.grid(column=0, row=15, columnspan=7, sticky='ew', pady=10)

vseperator1 = ttk.Separator(root, orient='vertical')
vseperator1.grid(column=0, row=0, rowspan=15, sticky='ns', padx=10)

vseperator2 = ttk.Separator(root, orient='vertical')
vseperator2.grid(column=3, row=0, rowspan=15, sticky='ns', padx=10)

vseperator3 = ttk.Separator(root, orient='vertical')
vseperator3.grid(column=6, row=0, rowspan=15, sticky='ns', padx=10)

#List of options for each variable

media_options = input_metadata_test['media']

strain_options = input_metadata_test['strain']

supp_options = input_metadata_test['supplement']

vector_options = input_metadata_test['vector']

gui_output = []
#gui_output is an open list to add metadata selections to

#Below are the dropdown menus for each metadata variable
# MEDIA SELECTION DROP DOWN #
label = Label(root, text = "Select Media(s) from Menu")
label.grid(column =1, row =2, sticky=tk.N+tk.S+tk.W+tk.E)   

media = tk.StringVar(root)

media.set("-Media-")

media_option = tk.OptionMenu(root, media, *media_options) #ENTER a command to reduce the selection process
media_option.grid(column =1, row =3, sticky=tk.N+tk.S+tk.W+tk.E)

#adding selections to a media list to reference later
media_list = set()
max_mediaitems = len(media_options)

def media_dropdown(*args):
    global dropdown
    dropdown = str(media.get())
    print(dropdown)

    if media.get() == media_options[0]:
        media_select = media_options[0]

    if media.get() == media_options[1]:
        media_select = media_options[1]

    #Repeat above if statement when adding another option to the metadata dictionary

    if len(media_list) < max_mediaitems:
        media_list.add(media_select)
    else:
        print("Max items reached")

    result_label.config(text=f"{media_list}", fg = "blue")
    result_label.grid(column =1, row=4, sticky=tk.N+tk.S+tk.W+tk.E)

def delete_mediaitems():
    media_select = media.get()
    if media_select in media_list:
        media_list.remove(media_select)
        print(f"Removed {media_select}")
    else:
        print(f"{media_select} not in list")
    
    result_label.config(text=f"{media_list}", fg="blue")
    result_label.grid(column=1, row=4, sticky=tk.N+tk.S+tk.W+tk.E)

#adding the selected media to the open list above
gui_output.append(media_list)

# link function to change dropdown
media_button = Button(root, text = "add", bg='#85e472', fg='black', command = media_dropdown)
media_button.grid(column = 2, row =3, sticky=tk.N+tk.S+tk.W+tk.E )

media_delete = Button(root, text = "delete", bg='#e47272', fg='black', command = delete_mediaitems)
media_delete.grid(column = 2, row =4, sticky=tk.N+tk.S+tk.W+tk.E )

####
####
####
####
####

# STRAIN SELECTION DROP DOWN MENU #
label = Label(root, text = "Select Strain(s) from Menu")
label.grid(column =4, row =2, sticky=tk.N+tk.S+tk.W+tk.E)

strain = tk.StringVar(root)

strain.set("-Strain-")

strain_option = tk.OptionMenu(root, strain, *strain_options)
strain_option.grid(column =4, row =3, sticky=tk.N+tk.S+tk.W+tk.E)

#adding strain selections to a list to reference later
strain_list = set()
max_strainitems = len(strain_options)

def strain_dropdown(*args):
    global dropdown
    dropdown = str(strain.get())
    print(dropdown)

    if strain.get() == strain_options[0]:
        strain_select = strain_options[0]

    if strain.get() == strain_options[1]:
        strain_select = strain_options[1]

    if strain.get() == strain_options[2]:
        strain_select = strain_options[2]

    if len(strain_list) < max_strainitems:
        strain_list.add(strain_select)
    else:
        print("Max items reached")

    sresult_label.config(text=f"{strain_list}", fg = "blue")
    sresult_label.grid(column =4, row=4, sticky=tk.N+tk.S+tk.W+tk.E)

def delete_strainitems():
    strain_select = strain.get()
    if strain_select in strain_list:
        strain_list.remove(strain_select)
        print(f"Removed {strain_select}")
    else:
        print(f"{strain_select} not in list")
    
    sresult_label.config(text=f"{strain_list}", fg = "blue")
    sresult_label.grid(column =4, row=4, sticky=tk.N+tk.S+tk.W+tk.E)
    
gui_output.append(strain_list)

strain_button = Button(root, text = "add", bg='#85e472', fg='black', command = strain_dropdown)
strain_button.grid(column = 5, row =3, sticky=tk.N+tk.S+tk.W+tk.E )

strain_delete = Button(root, text = "delete", bg='#e47272', fg='black', command = delete_strainitems)
strain_delete.grid(column = 5, row =4, sticky=tk.N+tk.S+tk.W+tk.E )

####
####
####
####
####

# VECTOR SELECTION DROP DOWN MENU #
label = Label(root, text = "Select Vector(s) from Menu")
label.grid(column =1, row =6, sticky=tk.N+tk.S+tk.W+tk.E)

vector = tk.StringVar(root)

vector.set("-Vector-")

vector_option = tk.OptionMenu(root, vector, *vector_options) #ENTER a command to reduce the selection process
vector_option.grid(column =1, row =7, sticky=tk.N+tk.S+tk.W+tk.E)

#adding selections to a media list to reference later
vector_list = set()
max_vectoritems = len(vector_options)

def vector_dropdown(*args):
    global dropdown
    dropdown = str(vector.get())
    print(dropdown)

    if vector.get() == vector_options[0]:
        vector_select = vector_options[0]

    if vector.get() == vector_options[1]:
        vector_select = vector_options[1]

    #Repeat above if statement when adding another option

    if len(vector_list) < max_vectoritems:
        vector_list.add(vector_select)
    else:
        print("Max items reached")

    vresult_label.config(text=f"{vector_list}", fg = "blue")
    vresult_label.grid(column =1, row=8, sticky=tk.N+tk.S+tk.W+tk.E)

def delete_vectoritems():
    vector_select = vector.get()
    if vector_select in vector_list:
        vector_list.remove(vector_select)
        print(f"Removed {vector_select}")
    else:
        print(f"{vector_select} not in list")

    vresult_label.config(text=f"{vector_list}", fg = "blue")
    vresult_label.grid(column =1, row=8, sticky=tk.N+tk.S+tk.W+tk.E)


gui_output.append(vector_list)

vector_button = Button(root, text = "add", bg='#85e472', fg='black', command = vector_dropdown)
vector_button.grid(column =2, row =7, sticky=tk.N+tk.S+tk.W+tk.E )

vector_delete = Button(root, text = "delete", bg='#e47272', fg='black', command = delete_vectoritems)
vector_delete.grid(column =2, row =8, sticky=tk.N+tk.S+tk.W+tk.E )

####
####
####
####
####

# SUPPLEMENT SELECTION DROP DOWN MENU #
label = Label(root, text = "Select Supplement(s) from Menu")
label.grid(column =4, row =6, sticky=tk.N+tk.S+tk.W+tk.E)

supp = tk.StringVar(root)

supp.set("-Supplement-")

supp_option = tk.OptionMenu(root, supp, *supp_options)
supp_option.grid(column =4, row =7, sticky=tk.N+tk.S+tk.W+tk.E)

#Dropdown selection loop to add the selected values to the overall temp list
supp_list = set()
max_suppitems = len(supp_options)

def supp_dropdown(*args):
    global dropdown
    dropdown = str(supp.get())
    print(dropdown)

    if supp.get() == supp_options[0]:
        supp_select = supp_options[0]

    if supp.get() == supp_options[1]:
        supp_select = supp_options[1]

    #Repeat above if statement when adding another option

    if len(supp_list) < max_suppitems:
        supp_list.add(supp_select)
    else:
        print("Max items reached")

    suresult_label.config(text=f"{supp_list}", fg = "blue")
    suresult_label.grid(column =4, row=8, sticky=tk.N+tk.S+tk.W+tk.E)

def delete_suppitems():
    supp_select = supp.get()
    if supp_select in supp_list:
        supp_list.remove(supp_select)
        print(f"Removed {supp_select}")
    else:
        print(f"{supp_select} not in list")
    
    suresult_label.config(text=f"{supp_list}", fg = "blue")
    suresult_label.grid(column =4, row=8, sticky=tk.N+tk.S+tk.W+tk.E)

gui_output.append(supp_list)

supp_button = Button(root, text = "add", bg='#85e472', fg='black', command = supp_dropdown)
supp_button.grid(column = 5, row =7, sticky=tk.N+tk.S+tk.W+tk.E )

supp_delete = Button(root, text = "delete", bg='#e47272', fg='black', command = delete_suppitems)
supp_delete.grid(column = 5, row =8, sticky=tk.N+tk.S+tk.W+tk.E )


num_copies = 0


def show_value(value):
    global num_copies
    num_copies = int(value)
    print(num_copies)


repli_scale = tk.Scale(root, from_=1, to=10, orient=HORIZONTAL, command=show_value)
repli_scale.grid(column = 1, row = 11, sticky = tk.N+tk.S+tk.W+tk.E)

repliresult_label = tk.Label(root, text = 'Select Number of Replicates', fg = "black")
repliresult_label.grid(column =1, row =10)

#Temp list button to check if the variables were added correctly
def temp_list1():
    temp_list.config(text=f'{gui_output}', fg = "blue")
    temp_list.grid(column = 4, row = 11, sticky =tk.N+tk.S+tk.W+tk.E)
    print(gui_output)


temp_listbutton = Button(root, text = "Check Selection List", command = temp_list1)
temp_listbutton.grid(column=4, row=10, sticky=tk.N+tk.S+tk.W+tk.E)
#End


#Below are label placeholders for the visual printing of the list values above
result_label = tk.Label(root, text = "", fg = "black")
result_label.grid(column =1, row =12)

sresult_label = tk.Label(root, text = "", fg = "black")
sresult_label.grid(column =1, row = 13)

suresult_label = tk.Label(root, text = "", fg = "black")
suresult_label.grid(column =1, row = 14)

vresult_label = tk.Label(root, text = "", fg = "black")
suresult_label.grid(column =1, row = 14)

temp_list = tk.Label(root)

label = tk.Label(root, text = " ")
label.grid()
#End label placeholder code

#Below is excel sheet generation code
def excel_sheet():
    try:
        product = list(itertools.product(*gui_output))
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Sample Design"
        titles = ['Sample Design ID', 
                  'Media ID', 
                  'Strain ID', 
                  'Vector ID', 
                  'Supplement ID']

#Populate the first row with the titles
        for col_num, title in enumerate(titles, start=1):
            cell = sheet1.cell(row=1, column=col_num, value=title)
            cell.font = Font(bold=True)

        global sample_design_id
        sample_design_id = 1

#Populating the excel sheet with generated data
        for row_num, item in enumerate(product, start=2):
    #Sample ID counter for first row
            sheet1.cell(row=row_num, column=1, 
                        value=f'SampleDesign{sample_design_id}')
            sample_design_id += 1

    #Continue to populate the rest of the row
            for col_num, value in enumerate(item, start=2):
                sheet1.cell(row=row_num, column=col_num, value=value)

        plate_96_wells_coordinates = [
        (1,1),(1,2),(1,3),(1,4),(1,5),(1,6),(1,7),(1,8),(1,9),(1,10),(1,11),(1,12),
        (2,1),(2,2),(2,3),(2,4),(2,5),(2,6),(2,7),(2,8),(2,9),(2,10),(2,11),(2,12),
        (3,1),(3,2),(3,3),(3,4),(3,5),(3,6),(3,7),(3,8),(3,9),(3,10),(3,11),(3,12),
        (4,1),(4,2),(4,3),(4,4),(4,5),(4,6),(4,7),(4,8),(4,9),(4,10),(4,11),(4,12),
        (5,1),(5,2),(5,3),(5,4),(5,5),(5,6),(5,7),(5,8),(5,9),(5,10),(5,11),(5,12),
        (6,1),(6,2),(6,3),(6,4),(6,5),(6,6),(6,7),(6,8),(6,9),(6,10),(6,11),(6,12),
        (7,1),(7,2),(7,3),(7,4),(7,5),(7,6),(7,7),(7,8),(7,9),(7,10),(7,11),(7,12),
        (8,1),(8,2),(8,3),(8,4),(8,5),(8,6),(8,7),(8,8),(8,9),(8,10),(8,11),(8,12)
        ]

        def generate_sample_names(num_designs):
            global sample_names
            sample_names = []
            for i in range(0, sample_design_id-1):
                sample_names.append(f'SampleDesign{i+1}')
            return sample_names
        
        num_designs = 4
        generate_sample_names(num_designs)

        replicates = num_copies

    # This creates/formats sheet 2 with the samples and their positions in a well plate
        sheet2 = wb.create_sheet(title="Sample")

        headers = ['Sample ID', 'Row', 'Column', 'Well ID', 'Sample Design ID']

        for col_num, header in enumerate(headers, start=1): 
            sheet2.cell(row=1, column=col_num, value=header)

        #Initializes the well count and sample ID count
        well_count = 0
        sample_id = 1

        for sd in sample_names:
            for r in range(replicates):
                    row_data = [
                        f'Sample{sample_id}',
                        plate_96_wells_coordinates[well_count][0],
                        plate_96_wells_coordinates[well_count][1],
                        f'Assay{+1}',
                        sd
                        ]
                    sheet2.append(row_data)
                    well_count += 1
                    sample_id += 1

# Auto fit column width code
        for col in sheet1.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet1.column_dimensions[column].width = adjusted_width

        for col in sheet2.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet2.column_dimensions[column].width = adjusted_width

        timestamp = datetime.now().strftime('%Y-%m-%d %H%M')
        filename = f'SampleDesign_{timestamp}.xlsx'

        wb.save(filename)

        os.startfile(filename)
    
    #Below is error handling code
    except Exception as e:
        error = tk.Tk()
        error.title("Error Window")
        error.withdraw()
        messagebox.showerror("Error Window", 
                             f"An error occurred during generation: {str(e)}\n\nPlease select a lower number of replicates and try again.")
        error.destroy()

excel_button = Button(root, text = "Generate Excel Sheet", 
                      font = ("Arial", 12, "bold"),
                      bg = "#c2bebe",
                      fg = "black",
                      borderwidth=1,
                      command = excel_sheet)
excel_button.grid(column = 4, row = 14, sticky=tk.N+tk.S+tk.W+tk.E)

root.mainloop()